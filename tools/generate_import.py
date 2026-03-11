"""
Script to generate product import Excel from LISTA PRODUCTOS SISTEMA.xlsx
Maps to product_template.xlsx format with brands, descriptions, and SAT codes.
"""
import openpyxl
from openpyxl import load_workbook

IVA = 1.16

# Barcode mapping: list of (keyword_set, ean) — all from Mexican retailers (750 prefix)
# Keywords must ALL appear in the product name (uppercase). More specific rules first.
BARCODE_RULES = [
    # Zote barra - solo presentaciones individuales (CAJA no tiene EAN propio)
    ({'ZOTE', 'ROSA',   '200G'},  '7501026005680'),
    ({'ZOTE', 'BLANCO', '200G'},  '7501026005381'),
    ({'ZOTE', 'ROSA',   '400G'},  '7501026005671'),
    ({'ZOTE', 'BLANCO', '400G'},  '7501026005374'),
    # Jabón Roma - solo individual (caja sin EAN propio)
    ({'ROMA', '1', 'KG'},          '7501026004602'),
    # (Roma 10KG bulk — sin EAN registrado)
    # Jabón Foca - solo individuales
    ({'FOCA', '1', 'KG'},          '7501026026546'),
    ({'FOCA', '5', 'KG'},          '7501026026505'),
    # Glade
    ({'GLADE', 'BAÑO'},            '7501032909860'),  # gel cono 170 g baño
    ({'GLADE', 'GEL'},             '7501032909860'),  # fallback gel
    ({'GLADE', 'CARRO'},           '7501032913030'),  # gel auto 70 g
]


def get_barcode(name: str) -> str:
    """Return EAN-13 for product if found, else empty string.
    CAJA/multi-pack presentations don't get a barcode (no registered case-level EAN).
    """
    tokens = set(name.upper().split())
    if 'CAJA' in tokens:
        return ''
    for keywords, ean in BARCODE_RULES:
        if all(kw in tokens for kw in keywords):
            return ean
    return ''


SOURCE = r'C:/Users/Manuel Alfaro/Documents/Desarrollos/Cleeny/LISTA PRODUCTOS SISTEMA.xlsx'
TEMPLATE = r'C:/Users/Manuel Alfaro/Documents/Desarrollos/Cleeny/product_template.xlsx'
OUTPUT = r'C:/repos/Cleeny/tools/productos_lista_importacion.xlsx'

# Brand rules by CODIGO (exact match, highest priority)
CODIGO_BRAND_MAP = {
    'AROFE':   'Ferrari',
    'AROCH':   'Chic',
    'ALMO':    'Almorol',
    'ALMOE':   'Almorol',
    'CLORO':   '',
    'CLOROE':  '',
    'DESIND':  '',
    'DESINDE': '',
    'DESCOC':  '',
    'DESCOCE': '',
    'DESMOT':  '',
    'DETNEU':  '',
    'COLPLUC': 'ColorPlus',
    'COLPLUO': 'ColorPlus',
    'ZOTPLUB': 'La Corona',
    'ZOTPLUR': 'La Corona',
    'GELANT':  '',
    'DELICO':  'Delman',
    'DELIPE':  'Delman',
    'ULTRAAZ': 'Ultrabrillo',
    'ULTRAV':  'Ultrabrillo',
    'LIMPIVI': '',
    'MULTIAZ': 'Multiclean',
    'MULTILA': 'Multiclean',
    'MULTILI': 'Multiclean',
    'MULTITIL':'Multiclean',
    'MULTIAE': 'Multiclean',
    'MULTIAZ': 'Multiclean',
    'ULTRAPC': 'Ultrapino',
    'ULTRAPL': 'Ultrapino',
    'ULTRAP':  'Ultrapino',
    'PREBLA':  '',
    'PREVADR': '',
    'SARRI':   '',
    'SUAAMA':  'Suaviplus',
    'SUAAZU':  'Suaviplus',
    'SUAENJ':  'Suaviplus',
    'BOT1L':   '',
    'POR20':   '',
    'POR5':    '',
}

# Brand rules by WORD in name (checks whole words only)
WORD_BRAND_MAP = {
    'ROMA':        'La Corona',
    'FOCA':        'La Corona',
    'ZOTE':        'La Corona',
    'GLADE':       'SC Johnson',
    'WISE':        'Wiese',
    'PERICO':      'Plásticos Italianos',
    'PERICOL':     'Plásticos Italianos',
    'PRACTIQUITO': 'Plásticos Italianos',
    'CUBASA':      'Cubasa',
    'TORINO':      'Plastibol',
    'ALMOROL':     'Almorol',
    'FERRARI':     'Ferrari',
    'CHIC':        'Chic',
    'DELMAN':      'Delman',
    'ULTRABRILLO': 'Ultrabrillo',
    'MULTICLEAN':  'Multiclean',
    'ULTRAPINO':   'Ultrapino',
    'SARRICIDA':   'Sarricida',
    'SUAVIPLUS':   'Suaviplus',
    'ZOTPLUS':     'La Corona',
    'COLORPLUS':   'ColorPlus',
}

# SAT code rules: (keyword_list, sat_code)
# More specific rules first
SAT_RULES = [
    (['JABON', 'ROMA'],              '47131811'),
    (['JABON', 'FOCA'],              '47131811'),
    (['JABON', 'ZOTE'],              '47131811'),
    (['JABON', 'UTIL'],              '47131811'),
    (['JABON', 'BARRA'],             '47131811'),
    (['ZOTPLUS'],                    '47131811'),
    (['SUAVIPLUS'],                  '47131811'),
    (['SUAENJ'],                     '47131811'),
    (['SUAAMA'],                     '47131811'),
    (['SUAAZU'],                     '47131811'),
    (['AROMATIZANTE'],               '47131812'),
    (['AROFE'],                      '47131812'),
    (['AROCH'],                      '47131812'),
    (['GELANT'],                     '47131700'),
    (['GLADE'],                      '47131812'),
    (['WISE'],                       '47131812'),
    (['FERRARI'],                    '47131812'),
    (['CHIC'],                       '47131812'),
    (['DELMAN'],                     '47131700'),
    (['DELICO'],                     '47131700'),
    (['DELIPE'],                     '47131700'),
    (['JABON', 'MANOS'],             '47131700'),
    (['ALMOROL'],                    '25172300'),
    (['ALMOE'],                      '25172300'),
    (['SARRICIDA'],                  '47131800'),
    (['SARRI'],                      '47131800'),
    (['ESCOBA'],                     '47121500'),
    (['CEPILLO'],                    '47121500'),
    (['BASTON'],                     '47121500'),
    (['PRACTIQUITO'],                '47121500'),
    (['PERICO'],                     '47121500'),
    (['CUBETA'],                     '47121500'),
    (['EMBUDO'],                     '47121500'),
    (['ATOMIZADOR'],                 '47121500'),
    (['BOLSA'],                      '47121500'),
    (['BOTELLA'],                    '47121500'),
    (['PORRON'],                     '47121500'),
    (['CLORO'],                      '47131805'),
    (['DESENGRASANTE'],              '47131805'),
    (['DESIND'],                     '47131805'),
    (['DESCOC'],                     '47131805'),
    (['DESMOT'],                     '47131805'),
    (['DESCOCE'],                    '47131805'),
    (['DESINDE'],                    '47131805'),
    (['LIMPIADOR'],                  '47131805'),
    (['LIMPIVI'],                    '47131805'),
    (['MULTICLEAN'],                 '47131805'),
    (['MULTILA'],                    '47131805'),
    (['MULTILI'],                    '47131805'),
    (['MULTIAZ'],                    '47131805'),
    (['MULTITIL'],                   '47131805'),
    (['MULTIAE'],                    '47131805'),
    (['ULTRAPINO'],                  '47131805'),
    (['ULTRAPC'],                    '47131805'),
    (['ULTRAPL'],                    '47131805'),
    (['ULTRAP'],                     '47131805'),
    (['ULTRABRILLO'],                '47131805'),
    (['ULTRAV'],                     '47131805'),
    (['PRELAVADOR'],                 '47131805'),
    (['PREBLA'],                     '47131805'),
    (['PREVADR'],                    '47131805'),
    (['DETNEU'],                     '47131805'),
    (['GEL'],                        '47131805'),
    (['COLORPLUS'],                  '47131805'),
    (['COLPLUC'],                    '47131805'),
    (['COLPLUO'],                    '47131805'),
    (['ZOTPLUB'],                    '47131811'),
    (['ZOTPLUR'],                    '47131811'),
    (['ZOTPLUR'],                    '47131805'),
]

DEFAULT_SAT = '47131800'  # General cleaning solutions


def get_brand(name: str, codigo: str = '') -> str:
    cod = (codigo or '').strip().upper()
    # 1. Exact codigo match
    if cod and cod in CODIGO_BRAND_MAP:
        val = CODIGO_BRAND_MAP[cod]
        return val if val else 'Cleeny'
    # 2. Whole-word match in name
    words = set(name.upper().split())
    for word, brand in WORD_BRAND_MAP.items():
        if word in words:
            return brand
    # 3. Partial codigo prefix match
    if cod:
        for key, brand in CODIGO_BRAND_MAP.items():
            if cod.startswith(key) and brand:
                return brand
    return 'Cleeny'


def get_sat_code(name: str, codigo: str = '') -> str:
    upper = name.upper()
    cod_upper = (codigo or '').upper()
    for keywords, code in SAT_RULES:
        if all(kw in upper or kw in cod_upper for kw in keywords):
            return code
    return DEFAULT_SAT


def generate_description(name: str, brand: str, unit: str) -> str:
    """Generate a meaningful Spanish product description from product name."""
    n = name.strip().upper()

    # Category detection
    if 'JABON' in n and 'MANOS' in n:
        cat = 'Jabón líquido para manos'
    elif 'JABON' in n and 'BARRA' in n:
        cat = 'Jabón en barra para lavandería'
    elif 'JABON' in n and 'CAJA' in n:
        cat = 'Jabón en presentación de caja'
    elif 'JABON' in n:
        cat = 'Jabón de lavandería'
    elif 'AROMATIZANTE' in n and 'GEL' in n:
        cat = 'Aromatizante ambiental en gel'
    elif 'AROMATIZANTE' in n and 'AEROSOL' in n:
        cat = 'Aromatizante ambiental en aerosol'
    elif 'AROMATIZANTE' in n or 'AROFE' in n or 'AROCH' in n:
        cat = 'Aromatizante ambiental'
    elif 'ATOMIZADOR' in n:
        cat = 'Atomizador/rociador de plástico'
    elif 'BASTON' in n:
        cat = 'Bastón de limpieza metálico'
    elif 'BOLSA' in n and 'CAMISETA' in n:
        cat = 'Bolsa camiseta de plástico'
    elif 'BOLSA' in n and 'BASURA' in n and 'ROLLO' in n:
        cat = 'Bolsa para basura en rollo'
    elif 'BOLSA' in n and 'BASURA' in n:
        cat = 'Bolsa para basura'
    elif 'CEPILLO' in n and 'PLANCHA' in n:
        cat = 'Cepillo tipo plancha para limpieza'
    elif 'CEPILLO' in n:
        cat = 'Cepillo de limpieza'
    elif 'CUBETA' in n and 'FLEXIBLE' in n:
        cat = 'Cubeta flexible de plástico'
    elif 'CUBETA' in n:
        cat = 'Cubeta de plástico'
    elif 'EMBUDO' in n:
        cat = 'Embudo de plástico'
    elif 'ESCOBA' in n and 'ARCO' in n:
        cat = 'Escoba tipo arco'
    elif 'ESCOBA' in n:
        cat = 'Escoba de limpieza'
    elif 'CLORO' in n or 'CLOROE' in n or 'CLORAE' in n:
        cat = 'Cloro blanqueador desinfectante'
    elif 'DESENGRASANTE' in n and 'MOTOR' in n or 'DESMOT' in n:
        cat = 'Desengrasante para motores'
    elif 'DESENGRASANTE' in n and ('COCINA' in n or 'DESCOC' in n or 'DESCOCE' in n):
        cat = 'Desengrasante para cocina'
    elif 'DESENGRASANTE' in n or 'DESIND' in n or 'DESINDE' in n:
        cat = 'Desengrasante industrial'
    elif 'GEL ANTIBACTERIAL' in n or 'GELANT' in n:
        cat = 'Gel antibacterial para manos'
    elif 'JABON' in n and 'MANOS' in n:
        cat = 'Jabón líquido para manos'
    elif 'LIMPIADOR' in n and 'MULTIUSOS' in n:
        cat = 'Limpiador multiusos en spray'
    elif 'LIMPIAVIDRIOS' in n or 'LIMPIVI' in n:
        cat = 'Limpiador para vidrios y ventanas'
    elif 'LIMPIADOR' in n:
        cat = 'Limpiador líquido'
    elif 'DETERGENTE' in n or 'DETNEU' in n:
        cat = 'Detergente líquido neutro'
    elif 'PRELAVADOR' in n or 'PREBLA' in n or 'PREVADR' in n:
        cat = 'Prelavador blanqueador para ropa'
    elif 'SUAVIPLUS' in n or 'SUAAMA' in n or 'SUAAZU' in n or 'SUAENJ' in n:
        cat = 'Suavizante textil para ropa'
    elif 'ZOTPLUS' in n or 'ZOTPLUB' in n or 'ZOTPLUR' in n:
        cat = 'Jabón líquido p/ropa base Zote'
    elif 'COLORPLUS' in n or 'COLPLUC' in n or 'COLPLUO' in n:
        cat = 'Jabón líquido p/ropa de color'
    elif 'ULTRAPINO' in n or 'ULTRAPC' in n or 'ULTRAPL' in n:
        cat = 'Limpiador de pino desinfectante'
    elif 'ULTRABRILLO' in n or 'ULTRAV' in n:
        cat = 'Lavatrastes lavavajillas líquido'
    elif 'ALMOROL' in n or 'ALMOE' in n:
        cat = 'Abrillantador para llantas y plásticos'
    elif 'SARRICIDA' in n or 'SARRI' in n:
        cat = 'Descalcificador para sanitarios'
    elif 'MULTICLEAN' in n or n.startswith('MULTI'):
        cat = 'Limpiador multiusos líquido'
    elif 'BOTELLA' in n:
        cat = 'Botella de plástico con tapa'
    elif 'PORRON' in n:
        cat = 'Porrón (contenedor) de plástico seminuevo'
    elif 'DELMAN' in n or 'DELICO' in n or 'DELIPE' in n:
        cat = 'Jabón líquido para manos'
    elif 'FERRARI' in n or 'AROFE' in n:
        cat = 'Aromatizante ambiental Ferrari'
    elif 'CHIC' in n or 'AROCH' in n:
        cat = 'Aromatizante ambiental Chic'
    else:
        cat = 'Producto de limpieza'

    # Size/format extraction
    size_tags = []
    for token in name.split():
        if token.endswith('KG') or token.endswith('G') and token[:-1].replace('.', '').isdigit():
            size_tags.append(token)
        elif token.endswith('ML') and token[:-2].replace('.', '').isdigit():
            size_tags.append(token)
        elif token.endswith('L') and token[:-1].replace('.', '').isdigit():
            size_tags.append(token + ' litros')
        elif token.isdigit() and int(token) > 0 and len(token) <= 3:
            pass  # skip bare numbers unless units follow

    # Color in name?
    color = ''
    for c in ['ROSA', 'BLANCO', 'AZUL', 'AMARILLO', 'VERDE', 'LIMON', 'LAVANDA', 'FRESCO']:
        if c in n:
            color_map = {
                'ROSA': 'rosa', 'BLANCO': 'blanco', 'AZUL': 'azul',
                'AMARILLO': 'amarillo', 'VERDE': 'verde', 'LIMON': 'limón',
                'LAVANDA': 'lavanda', 'FRESCO': 'azul fresco'
            }
            color = f', fragancia {color_map[c]}'
            break

    # Envasado?
    envasado = ' envasado' if 'ENVASADO' in n else ''

    # Unit label
    unit_label = ''
    if unit:
        u = str(unit).upper()
        if u == 'LT':
            unit_label = ', se vende por litro'
        elif u == 'PZ':
            unit_label = ', se vende por pieza'

    # Presentation (CAJA)?
    pres = ''
    if 'CAJA' in n:
        for token in n.split():
            if token.isdigit() and int(token) > 0:
                pass
        pres = ' (presentación en caja)'

    brand_part = f' marca {brand}' if brand else ''
    size_part = f' {" ".join(size_tags)}' if size_tags else ''

    return f'{cat}{brand_part}{size_part}{color}{envasado}{pres}{unit_label}.'


UNIT_MAP = {
    'PZ':  'PZA',
    'PZA': 'PZA',
    'CJ':  'PZA',   # caja → se vende por pieza/caja
    'LT':  'L',
    'L':   'L',
    'LPZ': 'L',     # litros/pieza → litros
    'KG':  'KG',
    'G':   'G',
    'ML':  'ML',
    'M':   'M',
    'CM':  'CM',
    'PAR': 'PAR',
    'JGO': 'JGO',
    'KIT': 'KIT',
}


def normalize_unit(unit) -> str:
    if not unit:
        return 'PZA'
    return UNIT_MAP.get(str(unit).strip().upper(), str(unit).strip().upper())


def remove_iva(price):
    if price is None:
        return None
    try:
        return round(float(price) / IVA, 4)
    except Exception:
        return None


def calc_discount_pct(precio_men, precio_may):
    try:
        men = float(precio_men)
        may = float(precio_may)
        if men > 0:
            return max(0, round((men - may) / men * 100, 2))
    except Exception:
        pass
    return 0


def write_row(ws, row_num, codigo, nombre, unit, costo, precio_men_raw, precio_may_raw, min_mayoreo):
    brand = get_brand(nombre, codigo)
    sat = get_sat_code(nombre, codigo)
    desc = generate_description(nombre, brand, unit)
    discount = calc_discount_pct(precio_men_raw, precio_may_raw)

    ws.cell(row=row_num, column=1).value = codigo if codigo else ''
    ws.cell(row=row_num, column=2).value = str(nombre).strip()
    ws.cell(row=row_num, column=3).value = brand
    ws.cell(row=row_num, column=4).value = desc
    ws.cell(row=row_num, column=5).value = get_barcode(nombre)          # Codigo Barras
    ws.cell(row=row_num, column=6).value = 1                            # Contenido
    ws.cell(row=row_num, column=7).value = normalize_unit(unit)
    ws.cell(row=row_num, column=8).value = remove_iva(costo)
    ws.cell(row=row_num, column=9).value = remove_iva(precio_men_raw)
    ws.cell(row=row_num, column=10).value = 'true'                      # Es Gravable
    ws.cell(row=row_num, column=11).value = 'true'                      # Esta Activo
    ws.cell(row=row_num, column=12).value = sat
    ws.cell(row=row_num, column=13).value = 'false'                     # Venta Parcial
    ws.cell(row=row_num, column=14).value = 'false'                     # Precio Personalizado
    ws.cell(row=row_num, column=15).value = 0                           # Cant Min Medio Mayoreo
    ws.cell(row=row_num, column=16).value = 0                           # Desc% Medio Mayoreo
    ws.cell(row=row_num, column=17).value = min_mayoreo
    ws.cell(row=row_num, column=18).value = discount


def main():
    src = load_workbook(SOURCE, data_only=True)
    tpl = load_workbook(TEMPLATE)
    ws_out = tpl['Products']

    # Clear example row
    for col in range(1, 19):
        ws_out.cell(row=2, column=col).value = None

    current_row = 2

    # --- JARCERIA ---
    # Cols (0-indexed): 0=NOMBRE,1=UNI,3=PU_prov,6=VTA_MEN,7=MAY1
    ws_j = src['JARCERIA']
    jarceria_count = 0
    for row in ws_j.iter_rows(min_row=6, values_only=True):
        nombre = row[0]
        if not nombre or not str(nombre).strip():
            continue
        nombre_str = str(nombre).strip()
        if nombre_str.upper().startswith('EN '):
            continue
        unit = row[1]
        costo = row[3]       # P.U. proveedor
        precio_men = row[6]  # VTA MEN
        precio_may = row[7]  # MAY 1
        write_row(ws_out, current_row, '', nombre_str, unit, costo, precio_men, precio_may, 6)
        current_row += 1
        jarceria_count += 1

    # --- PRODUCTO TERMINADO ---
    # Cols (0-indexed): 0=CODIGO,1=NOMBRE,2=UNI,4=Prov_MAY,5=VTA_MEN,6=MAY1_VTA
    ws_pt = src['PRODUCTO TERMINADO']
    pt_count = 0
    for row in ws_pt.iter_rows(min_row=6, values_only=True):
        nombre = row[1]
        if not nombre or not str(nombre).strip():
            continue
        nombre_str = str(nombre).strip()
        if nombre_str.upper().startswith('EN '):
            continue
        codigo = str(row[0]).strip() if row[0] else ''
        unit = row[2]
        costo = row[4]       # Prov MAY
        precio_men = row[5]  # VTA MEN
        precio_may = row[6]  # MAY 1 VTA
        write_row(ws_out, current_row, codigo, nombre_str, unit, costo, precio_men, precio_may, 20)
        current_row += 1
        pt_count += 1

    tpl.save(OUTPUT)

    print(f'Guardado: {OUTPUT}')
    print(f'JARCERIA: {jarceria_count} productos')
    print(f'PRODUCTO TERMINADO: {pt_count} productos')
    print(f'Total: {jarceria_count + pt_count} productos')
    print()
    print('=== Muestra (primeras 5 filas) ===')
    headers = ['Código','Nombre','Marca','Descripción','CodBar','Cont','Unidad','Costo','Precio',
               'Gravable','Activo','SAT','VentaParcial','PrecioCustom','MinMedMay','DescMedMay','MinMay','DescMay']
    print(' | '.join(f'{h[:8]}' for h in headers))
    for r in range(2, 7):
        vals = [str(ws_out.cell(row=r, column=c).value or '')[:10] for c in range(1, 19)]
        print(' | '.join(f'{v:<10}' for v in vals))


if __name__ == '__main__':
    main()
