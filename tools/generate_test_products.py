"""
Genera un archivo Excel de prueba para carga masiva de productos.
Incluye productos por unidad y a granel, con columnas de mayoreo.

Uso: python tools/generate_test_products.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def generate_test_products():
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # ── Encabezados base (deben coincidir con TemplateService.es.resx) ──
    base_headers = [
        "Código",                          # 1  - Code (opcional, auto-genera)
        "Nombre",                          # 2  - Name
        "Marca",                           # 3  - Brand
        "Descripción",                     # 4  - Description
        "Código de Barras",                # 5  - Barcode
        "Contenido",                       # 6  - Content
        "Código de Unidad de Medida",      # 7  - UnitMeasureCode
        "Costo",                           # 8  - Cost
        "Precio",                          # 9  - Price
        "Es Gravable",                     # 10 - IsTaxable
        "Está Activo",                     # 11 - IsActive
        "Código SAT de Producto/Servicio", # 12 - MexicoProductServiceCode
        "Permitir Venta Parcial",          # 13 - AllowPartialSale
        "Permitir Precio Personalizado",   # 14 - AllowCustomPricing
    ]

    # Encabezados de mayoreo (deben coincidir con prefijos "Cant. Min." y "Descuento %")
    wholesale_headers = [
        "Cant. Min. Medio Mayoreo",   # 15
        "Descuento % Medio Mayoreo",  # 16
        "Cant. Min. Mayoreo",         # 17
        "Descuento % Mayoreo",        # 18
    ]

    all_headers = base_headers + wholesale_headers

    # ── Estilos ──
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    wholesale_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    header_font = Font(bold=True)
    wholesale_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(bottom=Side(style="thin"))

    # Escribir encabezados
    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = wholesale_font if col_idx > len(base_headers) else header_font
        cell.fill = wholesale_fill if col_idx > len(base_headers) else header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # ── Datos de prueba ──
    # Formato por tupla:
    #   (Código, Nombre, Marca, Descripción, CódBarras, Contenido, Unidad,
    #    Costo, Precio, Gravable, Activo, CódigoSAT, VentaParcial, PrecioPersonalizado,
    #    CantMinMedioMayoreo, DescuentoMedioMayoreo, CantMinMayoreo, DescuentoMayoreo)
    #
    # Unidades de medida MX válidas (del UnitMeasureSeeder):
    #   PZA=Piezas, L=Litros, KG=Kilogramos, G=Gramos, ML=Mililitros,
    #   M=Metros, CM=Centímetros, MM=Milímetros, M2=Metros Cuadrados,
    #   PAR=Pares, JGO=Juegos, KIT=Kit

    products = [
        # ══════════════════════════════════════════════════════════════
        # CLORO Y BLANQUEADORES - Por unidad
        # ══════════════════════════════════════════════════════════════
        ("", "Cloro Concentrado 1L", "CleanPro", "Cloro concentrado para limpieza general, presentación 1 litro",
         "7501234567001", 1, "PZA", 25.00, 45.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),
        ("", "Cloro Concentrado 4L", "CleanPro", "Cloro concentrado para limpieza general, presentación 4 litros",
         "7501234567002", 1, "PZA", 75.00, 135.00, "true", "true", "14111705", "false", "false",
         6, 5, 12, 10),
        ("", "Cloro Gel 750ml", "CleanPro", "Cloro en gel para superficies, aroma fresco",
         "7501234567003", 1, "PZA", 22.00, 42.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),
        ("", "Blanqueador con Aroma Lavanda 1L", "FreshAir", "Blanqueador con fragancia lavanda",
         "7501234567004", 1, "PZA", 20.00, 38.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),
        ("", "Blanqueador con Aroma Pino 1L", "FreshAir", "Blanqueador con fragancia pino",
         "7501234567005", 1, "PZA", 20.00, 38.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),

        # ══════════════════════════════════════════════════════════════
        # DESINFECTANTES Y ANTIBACTERIALES - Por unidad
        # ══════════════════════════════════════════════════════════════
        ("", "Desinfectante Multiusos 750ml", "CleanPro", "Desinfectante multiusos aroma lavanda",
         "7501234567010", 1, "PZA", 20.00, 38.50, "true", "true", "14111705", "false", "false",
         10, 5, 20, 12),
        ("", "Desinfectante Multiusos 1L", "CleanPro", "Desinfectante multiusos aroma lavanda, presentación 1L",
         "7501234567011", 1, "PZA", 28.00, 52.00, "true", "true", "14111705", "false", "false",
         10, 5, 20, 12),
        ("", "Desinfectante Multiusos 4L", "CleanPro", "Desinfectante multiusos aroma lavanda, presentación 4L",
         "7501234567012", 1, "PZA", 85.00, 155.00, "true", "true", "14111705", "false", "false",
         6, 5, 12, 10),
        ("", "Desinfectante en Spray 400ml", "Hygiene Plus", "Desinfectante en aerosol para superficies",
         "7501234567013", 1, "PZA", 35.00, 65.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),
        ("", "Antibacterial para Manos 250ml", "Hygiene Plus", "Gel antibacterial con aloe vera",
         "7501234567014", 1, "PZA", 18.00, 35.00, "true", "true", "14111705", "false", "false",
         20, 8, 48, 15),
        ("", "Antibacterial para Manos 500ml", "Hygiene Plus", "Gel antibacterial con aloe vera, presentación grande",
         "7501234567015", 1, "PZA", 30.00, 58.00, "true", "true", "14111705", "false", "false",
         12, 8, 24, 15),
        ("", "Antibacterial para Manos 1L", "Hygiene Plus", "Gel antibacterial con aloe vera, uso familiar",
         "7501234567016", 1, "PZA", 50.00, 95.00, "true", "true", "14111705", "false", "false",
         10, 8, 20, 15),
        ("", "Toallitas Desinfectantes x80", "Hygiene Plus", "Toallitas húmedas desinfectantes, paquete de 80",
         "7501234567017", 1, "PZA", 32.00, 62.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),

        # ══════════════════════════════════════════════════════════════
        # JABONES - Por unidad
        # ══════════════════════════════════════════════════════════════
        ("", "Jabón Líquido para Manos 250ml", "Hygiene Plus", "Jabón antibacterial con aloe vera",
         "7501234567020", 1, "PZA", 12.00, 24.00, "true", "true", "14111705", "false", "false",
         24, 8, 48, 15),
        ("", "Jabón Líquido para Manos 500ml", "Hygiene Plus", "Jabón antibacterial con aloe vera, 500ml",
         "7501234567021", 1, "PZA", 16.00, 32.00, "true", "true", "14111705", "false", "false",
         15, 8, 30, 15),
        ("", "Jabón Líquido para Manos 1L", "Hygiene Plus", "Jabón antibacterial refill 1 litro",
         "7501234567022", 1, "PZA", 28.00, 55.00, "true", "true", "14111705", "false", "false",
         10, 8, 20, 15),
        ("", "Jabón en Barra Lavaplatos 350g", "ScrubbMax", "Jabón en barra para lavar trastes, aroma limón",
         "7501234567023", 1, "PZA", 8.00, 16.00, "true", "true", "14111705", "false", "false",
         30, 10, 60, 18),
        ("", "Jabón Lavaplatos Líquido 750ml", "ScrubbMax", "Jabón líquido concentrado para trastes",
         "7501234567024", 1, "PZA", 18.00, 35.00, "true", "true", "14111705", "false", "false",
         15, 8, 30, 15),
        ("", "Jabón Lavaplatos Líquido 1.5L", "ScrubbMax", "Jabón líquido concentrado para trastes, formato familiar",
         "7501234567025", 1, "PZA", 32.00, 60.00, "true", "true", "14111705", "false", "false",
         10, 8, 20, 12),

        # ══════════════════════════════════════════════════════════════
        # DETERGENTES Y SUAVIZANTES - Por unidad
        # ══════════════════════════════════════════════════════════════
        ("", "Detergente en Polvo 1kg", "CleanPro", "Detergente multiusos en polvo, aroma fresco",
         "7501234567030", 1, "PZA", 22.00, 42.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),
        ("", "Detergente en Polvo 5kg", "CleanPro", "Detergente multiusos en polvo, bolsa grande",
         "7501234567031", 1, "PZA", 85.00, 155.00, "true", "true", "14111705", "false", "false",
         6, 8, 12, 15),
        ("", "Detergente Líquido 1L", "CleanPro", "Detergente líquido concentrado para ropa",
         "7501234567032", 1, "PZA", 35.00, 65.00, "true", "true", "14111705", "false", "false",
         10, 5, 20, 10),
        ("", "Detergente Líquido 3L", "CleanPro", "Detergente líquido concentrado, formato familiar",
         "7501234567033", 1, "PZA", 80.00, 148.00, "true", "true", "14111705", "false", "false",
         6, 5, 12, 10),
        ("", "Suavizante de Telas 850ml", "FreshAir", "Suavizante concentrado aroma primavera",
         "7501234567034", 1, "PZA", 25.00, 48.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),
        ("", "Suavizante de Telas 3L", "FreshAir", "Suavizante concentrado, formato familiar",
         "7501234567035", 1, "PZA", 65.00, 120.00, "true", "true", "14111705", "false", "false",
         6, 5, 12, 10),

        # ══════════════════════════════════════════════════════════════
        # LIMPIADORES ESPECIALIZADOS - Por unidad
        # ══════════════════════════════════════════════════════════════
        ("", "Limpiador Multiusos Pino 1L", "PinolMax", "Limpiador con aroma natural de pino",
         "7501234567040", 1, "PZA", 15.00, 28.00, "true", "true", "14111705", "false", "false",
         15, 5, 30, 10),
        ("", "Limpiador Multiusos Pino 4L", "PinolMax", "Limpiador pino formato económico",
         "7501234567041", 1, "PZA", 45.00, 82.00, "true", "true", "14111705", "false", "false",
         6, 5, 12, 10),
        ("", "Limpiador de Vidrios 750ml", "CrystalClean", "Limpiador de cristales y espejos con spray",
         "7501234567042", 1, "PZA", 20.00, 38.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),
        ("", "Desengrasante Cocina 750ml", "QuimiClean", "Desengrasante de cocina, fórmula potente",
         "7501234567043", 1, "PZA", 25.00, 48.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),
        ("", "Limpiador de Baños 750ml", "Hygiene Plus", "Limpiador y desinfectante para baños",
         "7501234567044", 1, "PZA", 22.00, 42.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),
        ("", "Limpiador de Pisos Cerámicos 1L", "CrystalClean", "Limpiador especializado para pisos cerámicos",
         "7501234567045", 1, "PZA", 28.00, 52.00, "true", "true", "14111705", "false", "false",
         10, 5, 20, 10),
        ("", "Limpiador de Acero Inoxidable 500ml", "CrystalClean", "Limpiador y abrillantador de acero inoxidable",
         "7501234567046", 1, "PZA", 35.00, 68.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),
        ("", "Destapa Caños Líquido 500ml", "QuimiClean", "Destapador de cañerías fórmula concentrada",
         "7501234567047", 1, "PZA", 30.00, 58.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),
        ("", "Ácido Muriático 1L", "QuimiClean", "Ácido muriático para limpieza pesada de pisos y baños",
         "7501234567048", 1, "PZA", 12.00, 22.00, "true", "true", "14111705", "false", "false",
         20, 8, 48, 15),
        ("", "Sarricida 500ml", "QuimiClean", "Removedor de sarro para baños y azulejos",
         "7501234567049", 1, "PZA", 25.00, 48.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),

        # ══════════════════════════════════════════════════════════════
        # AROMATIZANTES - Por unidad
        # ══════════════════════════════════════════════════════════════
        ("", "Aromatizante Spray Primavera 400ml", "FreshAir", "Aromatizante ambiental aroma primavera",
         "7501234567050", 1, "PZA", 28.00, 55.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),
        ("", "Aromatizante Spray Lavanda 400ml", "FreshAir", "Aromatizante ambiental aroma lavanda",
         "7501234567051", 1, "PZA", 28.00, 55.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),
        ("", "Aromatizante Spray Cítrico 400ml", "FreshAir", "Aromatizante ambiental aroma cítrico",
         "7501234567052", 1, "PZA", 28.00, 55.00, "true", "true", "14111705", "false", "false",
         12, 5, 24, 10),
        ("", "Pastilla Sanitaria WC Pino", "FreshAir", "Pastilla desodorante para tanque de WC",
         "7501234567053", 1, "PZA", 8.00, 18.00, "true", "true", "14111705", "false", "false",
         30, 10, 60, 18),
        ("", "Pastilla Sanitaria WC Lavanda", "FreshAir", "Pastilla desodorante para tanque de WC aroma lavanda",
         "7501234567054", 1, "PZA", 8.00, 18.00, "true", "true", "14111705", "false", "false",
         30, 10, 60, 18),

        # ══════════════════════════════════════════════════════════════
        # PAPEL Y DESECHABLES - Por unidad
        # ══════════════════════════════════════════════════════════════
        ("", "Toallas Desechables Rollo x100", "PaperClean", "Toallas de papel absorbente, rollo de 100 hojas",
         "7501234567060", 1, "PZA", 15.00, 28.00, "true", "true", "14111704", "false", "false",
         20, 5, 50, 10),
        ("", "Toallas Desechables Rollo x250", "PaperClean", "Toallas de papel absorbente, rollo industrial",
         "7501234567061", 1, "PZA", 35.00, 65.00, "true", "true", "14111704", "false", "false",
         10, 5, 20, 10),
        ("", "Papel Higiénico Pack x4", "PaperClean", "Papel higiénico doble hoja, 4 rollos",
         "7501234567062", 1, "PZA", 18.00, 35.00, "true", "true", "14111704", "false", "false",
         15, 5, 30, 10),
        ("", "Papel Higiénico Pack x12", "PaperClean", "Papel higiénico doble hoja, 12 rollos",
         "7501234567063", 1, "PZA", 48.00, 90.00, "true", "true", "14111704", "false", "false",
         6, 8, 12, 15),
        ("", "Servilletas Pack x500", "PaperClean", "Servilletas de papel blanco, paquete de 500",
         "7501234567064", 1, "PZA", 25.00, 48.00, "true", "true", "14111704", "false", "false",
         12, 5, 24, 10),
        ("", "Bolsa de Basura 60x90 Paq x25", "TrashPro", "Bolsas negras para basura mediana",
         "7501234567065", 1, "PZA", 18.00, 35.00, "true", "true", "14111704", "false", "false",
         20, 8, 50, 15),
        ("", "Bolsa de Basura 90x120 Paq x10", "TrashPro", "Bolsas negras extra resistentes para basura grande",
         "7501234567066", 1, "PZA", 22.00, 42.00, "true", "true", "14111704", "false", "false",
         20, 8, 50, 15),
        ("", "Bolsa de Basura 90x120 Paq x25", "TrashPro", "Bolsas negras extra resistentes, paquete grande",
         "7501234567067", 1, "PZA", 48.00, 90.00, "true", "true", "14111704", "false", "false",
         10, 8, 20, 15),

        # ══════════════════════════════════════════════════════════════
        # UTENSILIOS DE LIMPIEZA - Por unidad
        # ══════════════════════════════════════════════════════════════
        ("", "Esponja Multiusos Pack x3", "ScrubbMax", "Esponjas doble uso, paquete de 3 piezas",
         "7501234567070", 1, "PZA", 10.00, 22.50, "true", "true", "14111706", "false", "false",
         24, 10, 48, 18),
        ("", "Esponja Multiusos Pack x6", "ScrubbMax", "Esponjas doble uso, paquete de 6 piezas",
         "7501234567071", 1, "PZA", 18.00, 38.00, "true", "true", "14111706", "false", "false",
         15, 10, 30, 18),
        ("", "Fibra Verde para Trastes Pack x3", "ScrubbMax", "Fibras abrasivas para limpieza de cocina",
         "7501234567072", 1, "PZA", 5.00, 12.00, "true", "true", "14111706", "false", "false",
         30, 10, 60, 18),
        ("", "Fibra Verde para Trastes Pack x10", "ScrubbMax", "Fibras abrasivas, paquete económico",
         "7501234567073", 1, "PZA", 14.00, 30.00, "true", "true", "14111706", "false", "false",
         15, 10, 30, 18),
        ("", "Estropajo de Acero Pack x3", "ScrubbMax", "Estropajo de acero inoxidable para ollas y sartenes",
         "7501234567074", 1, "PZA", 8.00, 18.00, "true", "true", "14111706", "false", "false",
         20, 10, 40, 18),
        ("", "Escoba de Plástico", "PlastiHome", "Escoba con cerdas de plástico resistente",
         "7501234567075", 1, "PZA", 25.00, 48.00, "true", "true", "14111706", "false", "false",
         "", "", "", ""),
        ("", "Escoba de Mijo Natural", "PlastiHome", "Escoba de fibra natural para exteriores",
         "7501234567076", 1, "PZA", 30.00, 55.00, "true", "true", "14111706", "false", "false",
         "", "", "", ""),
        ("", "Recogedor de Basura", "PlastiHome", "Recogedor de plástico con borde de goma",
         "7501234567077", 1, "PZA", 15.00, 28.00, "true", "true", "14111706", "false", "false",
         "", "", "", ""),
        ("", "Trapeador de Microfibra", "PlastiHome", "Trapeador con cabeza de microfibra intercambiable",
         "7501234567078", 1, "PZA", 38.00, 75.00, "true", "true", "14111706", "false", "false",
         "", "", "", ""),
        ("", "Trapeador de Algodón", "PlastiHome", "Trapeador de algodón con bastón de aluminio",
         "7501234567079", 1, "PZA", 28.00, 52.00, "true", "true", "14111706", "false", "false",
         "", "", "", ""),
        ("", "Repuesto Trapeador Microfibra", "PlastiHome", "Cabeza de repuesto para trapeador de microfibra",
         "7501234567080", 1, "PZA", 20.00, 38.00, "true", "true", "14111706", "false", "false",
         "", "", "", ""),
        ("", "Cubeta Plástica 12L", "PlastiHome", "Cubeta de plástico resistente con asa metálica",
         "7501234567081", 1, "PZA", 18.00, 35.00, "true", "true", "14111706", "false", "false",
         10, 5, 25, 10),
        ("", "Cubeta Plástica 20L", "PlastiHome", "Cubeta plástica uso industrial",
         "7501234567082", 1, "PZA", 28.00, 52.00, "true", "true", "14111706", "false", "false",
         10, 5, 25, 10),
        ("", "Atomizador Vacío 500ml", "PlastiHome", "Botella con atomizador reutilizable",
         "7501234567083", 1, "PZA", 10.00, 22.00, "true", "true", "14111706", "false", "false",
         20, 10, 50, 18),
        ("", "Cepillo para WC con Base", "ScrubbMax", "Cepillo para inodoro con soporte plástico",
         "7501234567084", 1, "PZA", 18.00, 35.00, "true", "true", "14111706", "false", "false",
         "", "", "", ""),
        ("", "Jalador para Pisos 40cm", "PlastiHome", "Jalador de agua para pisos con bastón",
         "7501234567085", 1, "PZA", 35.00, 65.00, "true", "true", "14111706", "false", "false",
         "", "", "", ""),

        # ══════════════════════════════════════════════════════════════
        # GUANTES Y PROTECCIÓN - Por unidad
        # ══════════════════════════════════════════════════════════════
        ("", "Guantes de Látex Caja x100 Ch", "SafeHands", "Guantes desechables talla chica, caja 100 piezas",
         "7501234567090", 1, "PZA", 90.00, 175.00, "true", "true", "14111703", "false", "false",
         5, 5, 10, 12),
        ("", "Guantes de Látex Caja x100 Med", "SafeHands", "Guantes desechables talla mediana, caja 100 piezas",
         "7501234567091", 1, "PZA", 95.00, 185.00, "true", "true", "14111703", "false", "false",
         5, 5, 10, 12),
        ("", "Guantes de Látex Caja x100 Gde", "SafeHands", "Guantes desechables talla grande, caja 100 piezas",
         "7501234567092", 1, "PZA", 95.00, 185.00, "true", "true", "14111703", "false", "false",
         5, 5, 10, 12),
        ("", "Guantes de Hule Amarillo Par", "SafeHands", "Guantes reutilizables para limpieza doméstica",
         "7501234567093", 1, "PAR", 12.00, 25.00, "true", "true", "14111703", "false", "false",
         20, 10, 50, 18),
        ("", "Cubrebocas Tricapa Caja x50", "SafeHands", "Cubrebocas desechable tricapa con elástico",
         "7501234567094", 1, "PZA", 35.00, 65.00, "true", "true", "14111703", "false", "false",
         10, 5, 20, 10),

        # ══════════════════════════════════════════════════════════════
        # KITS Y JUEGOS DE LIMPIEZA - Por unidad
        # ══════════════════════════════════════════════════════════════
        ("", "Kit Limpieza Básico Hogar", "CleanPro", "Kit con cloro 1L, desinfectante 750ml, esponja x3",
         "7501234567095", 1, "KIT", 55.00, 99.00, "true", "true", "14111705", "false", "false",
         6, 8, 12, 15),
        ("", "Kit Limpieza de Baño Completo", "Hygiene Plus", "Kit con limpiador baños, desinfectante, cepillo WC",
         "7501234567096", 1, "KIT", 65.00, 120.00, "true", "true", "14111705", "false", "false",
         6, 8, 12, 15),
        ("", "Juego de Escoba y Recogedor", "PlastiHome", "Set escoba plástica con recogedor y soporte",
         "7501234567097", 1, "JGO", 38.00, 72.00, "true", "true", "14111706", "false", "false",
         "", "", "", ""),

        # ══════════════════════════════════════════════════════════════
        # PRODUCTOS A GRANEL (L) - Líquidos, permiten venta parcial
        # ══════════════════════════════════════════════════════════════
        ("", "Cloro Industrial a Granel", "QuimiClean", "Cloro industrial al 13%, venta por litro",
         "", 1, "L", 8.00, 18.00, "true", "true", "14111705", "true", "true",
         20, 8, 50, 15),
        ("", "Cloro Regular a Granel", "QuimiClean", "Cloro al 6% para uso doméstico, venta por litro",
         "", 1, "L", 5.00, 12.00, "true", "true", "14111705", "true", "true",
         20, 8, 50, 15),
        ("", "Desengrasante Industrial a Granel", "QuimiClean", "Desengrasante concentrado, venta por litro",
         "", 1, "L", 12.00, 25.00, "true", "true", "14111705", "true", "true",
         15, 5, 40, 12),
        ("", "Jabón Líquido para Manos a Granel", "Hygiene Plus", "Jabón antibacterial a granel, venta por litro",
         "", 1, "L", 10.00, 22.00, "true", "true", "14111705", "true", "true",
         20, 8, 50, 15),
        ("", "Jabón Lavaplatos a Granel", "ScrubbMax", "Jabón líquido para trastes concentrado, venta por litro",
         "", 1, "L", 12.00, 24.00, "true", "true", "14111705", "true", "true",
         20, 8, 50, 15),
        ("", "Suavizante de Telas a Granel", "FreshAir", "Suavizante concentrado aroma floral, venta por litro",
         "", 1, "L", 9.00, 20.00, "true", "true", "14111705", "true", "true",
         15, 5, 30, 10),
        ("", "Pinol a Granel", "PinolMax", "Limpiador multiusos de pino, venta por litro",
         "", 1, "L", 7.00, 16.00, "true", "true", "14111705", "true", "true",
         20, 5, 50, 12),
        ("", "Limpiador de Pisos a Granel", "CleanPro", "Limpiador para pisos aroma floral, venta por litro",
         "", 1, "L", 6.00, 14.00, "true", "true", "14111705", "true", "true",
         20, 5, 50, 12),
        ("", "Ácido Muriático a Granel", "QuimiClean", "Ácido muriático para limpieza pesada, venta por litro",
         "", 1, "L", 5.00, 12.00, "true", "true", "14111705", "true", "true",
         20, 10, 50, 18),
        ("", "Aromatizante Líquido a Granel", "FreshAir", "Aromatizante concentrado multiusos, venta por litro",
         "", 1, "L", 8.00, 18.00, "true", "true", "14111705", "true", "true",
         15, 5, 40, 12),
        ("", "Detergente Líquido a Granel", "CleanPro", "Detergente líquido para ropa, venta por litro",
         "", 1, "L", 11.00, 22.00, "true", "true", "14111705", "true", "true",
         20, 8, 50, 15),

        # ══════════════════════════════════════════════════════════════
        # PRODUCTOS A GRANEL (KG) - Sólidos, permiten venta parcial
        # ══════════════════════════════════════════════════════════════
        ("", "Detergente en Polvo a Granel", "CleanPro", "Detergente multiusos en polvo, venta por kilo",
         "", 1, "KG", 18.00, 35.00, "true", "true", "14111705", "true", "true",
         10, 8, 25, 15),
        ("", "Bicarbonato de Sodio a Granel", "NaturClean", "Bicarbonato de sodio grado limpieza, venta por kilo",
         "", 1, "KG", 14.00, 28.00, "true", "true", "14111705", "true", "false",
         10, 5, 25, 12),
        ("", "Sosa Cáustica a Granel", "QuimiClean", "Hidróxido de sodio en escamas, venta por kilo",
         "", 1, "KG", 22.00, 45.00, "true", "true", "14111705", "true", "false",
         5, 5, 10, 10),
        ("", "Sal Industrial a Granel", "QuimiClean", "Sal gruesa para uso industrial y limpieza, venta por kilo",
         "", 1, "KG", 4.00, 10.00, "true", "true", "14111705", "true", "false",
         20, 8, 50, 15),
    ]

    # Escribir datos
    for row_idx, product in enumerate(products, 2):
        for col_idx, value in enumerate(product, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Formato texto para código de barras y código SAT
            if col_idx in (5, 12) and value:
                cell.number_format = "@"

    # ── Ajustar anchos de columna ──
    col_widths = {
        1: 12,   # Código
        2: 40,   # Nombre
        3: 16,   # Marca
        4: 55,   # Descripción
        5: 18,   # Código de Barras
        6: 12,   # Contenido
        7: 28,   # Unidad de Medida
        8: 12,   # Costo
        9: 12,   # Precio
        10: 14,  # Es Gravable
        11: 14,  # Está Activo
        12: 30,  # Código SAT
        13: 22,  # Venta Parcial
        14: 28,  # Precio Personalizado
        15: 26,  # Cant. Min. Medio Mayoreo
        16: 26,  # Descuento % Medio Mayoreo
        17: 22,  # Cant. Min. Mayoreo
        18: 22,  # Descuento % Mayoreo
    }

    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Altura de fila de encabezado
    ws.row_dimensions[1].height = 30

    # ── Guardar ──
    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "docs")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "productos_prueba_carga_masiva.xlsx")
    wb.save(output_path)

    # Estadísticas
    pza_count = sum(1 for p in products if p[6] == "PZA")
    kit_jgo = sum(1 for p in products if p[6] in ("KIT", "JGO", "PAR"))
    granel_l = sum(1 for p in products if p[6] == "L")
    granel_kg = sum(1 for p in products if p[6] == "KG")
    con_mayoreo = sum(1 for p in products if p[14] != "")
    sin_mayoreo = sum(1 for p in products if p[14] == "")

    print(f"Archivo generado: {output_path}")
    print(f"Total productos: {len(products)}")
    print(f"  - Por unidad (PZA):      {pza_count}")
    print(f"  - Kits/Juegos/Pares:     {kit_jgo}")
    print(f"  - A granel líquido (L):  {granel_l}")
    print(f"  - A granel sólido (KG):  {granel_kg}")
    print(f"  - Con mayoreo:           {con_mayoreo}")
    print(f"  - Sin mayoreo:           {sin_mayoreo}")

if __name__ == "__main__":
    generate_test_products()
